"""
||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
Author: Anupam Manna
Email ID: am7059141480@gmail.com
Mobile No: +91 7059141480
Date: October 2, 2023
Description: IREPS tenders automation.
||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
"""

import gc
import json
import os
import subprocess
import sys
import re
from openpyxl import Workbook, load_workbook
import time
# import tempfile
# import shutil
# import logging
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
# from email.mime.application import MIMEApplication
from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.chrome.options import Options
# import shutil
# import urllib.request
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from dateutil.relativedelta import relativedelta
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException
# from urllib.parse import urlparse
import requests
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import Select
import xlsxwriter
import math
import pdfplumber
import platform
import socket
import tempfile
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from chrome_updater import ChromeUpdater
from Program_Files.scraping_library import check_internet_connection
# from Program_Files.scraping_library import get_folder_size_in_mb
# from Program_Files.scraping_library import delete_empty_folders
from Program_Files.scraping_library import delete_folder
from Program_Files.scraping_library import packaging
from Program_Files.scraping_library import create_folder_if_not_exists
from Program_Files.scraping_library import is_android_device_connected
# from Program_Files.scraping_library import send_email
from Program_Files.scraping_library import countdown_timer
from Program_Files.scraping_library import delete_xlsx_files
from Program_Files.scraping_library import no_adb_mail
from Program_Files.scraping_library import skip_zones
from Program_Files.scraping_library import get_current_device_serial
from Program_Files.captcha_solver import predict_captcha

import base64
import io
from PIL import Image
# # =======================
# # chrome_updater
# # =======================
# # Create an instance of ChromeUpdater
# updater = ChromeUpdater()

# # Step 1: Open Chrome with Selenium
# driver = updater.relaunch_chrome_with_selenium()

# # Step 2: Close Chrome to prepare for the update
# time.sleep(5)  # Simulate some activity
# updater.close_chrome(driver)

# # Step 3: Download and install the latest version of Chrome
# updater.download_chrome_installer()
# updater.install_chrome()

# # Step 4: Relaunch Chrome after the update
# driver = updater.relaunch_chrome_with_selenium()




#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>># Global Variables
# Determine script directory
script_path = os.path.abspath(sys.argv[0])  # Absolute path of the executable
script_dir_path = os.path.dirname(script_path)

# Define directories and file paths relative to the script directory
# input_files_dir = os.path.join(script_dir_path, "Input_Files")
output_files_dir = os.path.join(script_dir_path, "Output_Files")
program_files_dir = os.path.join(script_dir_path, "Program_Files")


# Join the script path with the source path
temp_dir_path = os.path.join(program_files_dir, "ireps_temp")
# print(temp_dir_path)

# email_file_path = os.path.join(input_files_dir, "emailid_list.txt")
# print(email_file_path)
# sheet_names_file_path = os.path.join(input_files_dir,"sheet_names.txt")
# print(sheet_names_file_path)
script_log_file_path = os.path.join(program_files_dir, "script_log.log" )
# print(script_log_file_path)
config_file_path = os.path.join(program_files_dir, "Configration.json")
# print(config_file_path)
org_file_path = os.path.join(program_files_dir, "Organization_list.txt")
# print(org_file_path)
ireps_data = os.path.join(program_files_dir, "ireps_data.pkl")
# print(ireps_data)
# file_to_save_path = os.path.join(input_files_dir, "send_mail_log.txt")
# print(file_to_save_path)
tender_pdf_file_path = os.path.join(temp_dir_path, "tender.pdf")
# print(tender_pdf_file_path)

DEFAULT_WAIT_SECONDS = 20
SHORT_WAIT_SECONDS = 5
PDF_DOWNLOAD_TIMEOUT = 30
PDF_DOWNLOAD_RETRIES = 5
PDF_DOWNLOAD_CHUNK_SIZE = 1024 * 1024
CHROMEDRIVER_INSTALL_LOCK = threading.Lock()
OTP_REFRESH_LOCK = threading.Lock()


#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>># Global Variables

# List to store the organization
organizations = []

# Read data from the text file
with open(org_file_path, 'r') as file:
    lines = file.readlines()

for line in lines:
    # Check if the line starts with '#' and ignore those lines
    if not line.startswith('#'):
        parts = line.strip().split(': ')
        if len(parts) == 2:
            number = parts[0]
            name = parts[1]
            organizations.append((number, name))




# Load the JSON data from the file
with open(config_file_path, 'r') as file:
    data = json.load(file)
# Extract values from the JSON data
browser = data['browser']
adb_value = data['adb_device']
captcha_manual_input = data["captcha_manual_input"]
adb_device_ip = data['adb_device_ip']
mobile_no = data.get('mobile_no')
dump_location = data.get('dump_location')
excel_file_path = data.get('excel_file_path')
# otp_file_location = data['otp_file_location']
notification_emailids = data['notification_emailids']
receiver_emailids = data['receiver_emailids']
max_org_workers = data.get('max_org_workers', 1)
max_zone_workers = data.get('max_zone_workers', 1)
INDIAN_RAILWAY_ORG_NUMBER = "01"
# print(notification_emailids)
sender_email_id = data['sender_email_id']
sender_email_password = data['sender_email_password']
# time.sleep(1000)


# Edit the value of "signal_ireps"
data["signal_ireps"] = "FALSE"
data["signal_datelog"] = "FALSE"
# Write the modified data back to the file
with open(config_file_path, 'w') as file:
    json.dump(data, file, indent=4)

#----------------------------------------------------------------------------------------------- Extension to global variable






def retry_action(action, retries=3, delay=2, description="operation"):
    """Run a Selenium or file operation with bounded retries."""
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return action()
        except Exception as exc:
            last_error = exc
            print(f"{description} failed on attempt {attempt}/{retries}: {exc}")
            if attempt < retries:
                time.sleep(delay)
    raise last_error



def close_extra_windows(driver, initial_handle):
    """Close every browser window except the initial results window."""
    for window_handle in list(driver.window_handles):
        if window_handle == initial_handle:
            continue
        try:
            driver.switch_to.window(window_handle)
            driver.close()
        except Exception as exc:
            print(f"Unable to close extra browser window: {exc}")
    driver.switch_to.window(initial_handle)



def wait_for_new_window(driver, current_handles, timeout=DEFAULT_WAIT_SECONDS):
    """Wait until Selenium opens a new window and return the new handle."""
    WebDriverWait(driver, timeout).until(lambda d: len(d.window_handles) > len(current_handles))
    new_handles = [handle for handle in driver.window_handles if handle not in current_handles]
    if not new_handles:
        raise TimeoutException("New browser window did not open")
    return new_handles[-1]



def safe_click(driver, locator, timeout=DEFAULT_WAIT_SECONDS):
    """Wait for an element to be clickable, then click it."""
    element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(locator))
    element.click()
    return element



# Load dump location from JSON file
def load_otp():
    try:
        with open(config_file_path, "r") as file:
            data = json.load(file)
            return data.get("otp")
    except FileNotFoundError:
        return None





# Load dump location from JSON file
def load_otp_date():
    try:
        with open(config_file_path, "r") as file:
            data = json.load(file)
            return data.get("otp_date")
    except FileNotFoundError:
        return None






def is_otp_valid():
    # Get the current date
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    otp_date = load_otp_date()
    otp = load_otp()
    # Check if the dates match
    if current_date == otp_date: # date
        print(f"The OTP for today is: {otp}")
        return True
    else:
        print("No atching OTP found for today's date.")
        return False




def is_under_maintenance(driver, url):
    # Check if the page contains the specified text
    if "Module under maintenance" in driver.page_source:
        print(f"{url}  -  Module under maintenance")
    else:
        print(f"{url} link is Accessible ")
    return driver




# def get_verification(driver):
#     if captcha_manual_input == 1:

#          # blue #3BB9FF & red #640A0A
#     # Find all img elements inside the span with id "verimage" using XPath
#     img_tags = driver.find_elements(By.XPATH, "//span[@id='verimage']//img")

#     # Iterate through the img elements
#     for img_tag in img_tags:
#         src = img_tag.get_attribute("src")
#         if "Captcha.jpg?r=" in src:
#             captcha_chars = src.split('Captcha.jpg?r=')[1][:6]
#             # print("Image Source:", src)
#             print("verification Code: ", captcha_chars)
#     return driver, captcha_chars


def get_verification(driver):
    captcha_chars = None

    if captcha_manual_input == 1:
        # Locate the captcha <img>
        img_element = driver.find_element(By.ID, "imgCaptcha")
        src = img_element.get_attribute("src")

        if src.startswith("data:image"):
            # Base64 decode
            img_data = src.split(",")[1]
            image = Image.open(io.BytesIO(base64.b64decode(img_data)))
            image.show()  # Opens the image in default viewer

        # Ask user input
        captcha_chars = input("Enter CAPTCHA from the image: ").strip()

    else:
        # Step 1: Get captcha image element
        img_element = driver.find_element(By.ID, "imgCaptcha")
        src = img_element.get_attribute("src")

        os.makedirs(temp_dir_path, exist_ok=True)
        with tempfile.NamedTemporaryFile(suffix=".png", dir=temp_dir_path, delete=False) as temp_image:
            test_image = temp_image.name
            if src.startswith("data:image"):  # If base64 encoded
                header, encoded = src.split(",", 1)
                temp_image.write(base64.b64decode(encoded))
            else:  # If src is a URL
                response = requests.get(src, timeout=PDF_DOWNLOAD_TIMEOUT)
                response.raise_for_status()
                temp_image.write(response.content)

        try:
            predicted_text = predict_captcha(os.path.join(program_files_dir, "captcha_model.pth"), test_image)
            print(f"Predicted text: {predicted_text}")
            captcha_chars = predicted_text.strip()
        finally:
            if os.path.exists(test_image):
                os.remove(test_image)


    return driver, captcha_chars





def login(driver, mobile_no):
    """
    Log in to a website using the provided WebDriver instance and mobile number.

    Args:
        driver (WebDriver): Selenium WebDriver instance for browser.
        mobile_no (str): Mobile number for login.

    Returns:
        WebDriver: Updated WebDriver instance after login, or None if login fails.
    """

    # Attempt to refresh the page up to 3 times if a timeout occurs
    for _ in range(3):
        try:
            driver.refresh()
            break
        except TimeoutException:
            print("Timeout occurred. Retrying refresh...")
            time.sleep(2)
    else:
        raise TimeoutException("Exceeded maximum retries. Unable to refresh.")

    time.sleep(3)  # Short delay before starting login

    # Attempt the login process up to 3 times
    for attempt in range(100):
        try:
            # Accept alert if present, then proceed
            Alert(driver).accept()
        except NoAlertPresentException:
            pass  # Ignore if no alert

        # Get verification code and OTP, and fill in login details
        driver, ver_code = get_verification(driver)
        if ver_code is None:
            return None  # Return None if verification retrieval fails

        otp = load_otp()
        driver.execute_script(f"document.getElementById('mobileNo').value='{mobile_no}'")
        driver.execute_script(f"document.getElementById('verification').value='{ver_code}'")
        driver.execute_script(f"document.getElementById('otp').value='{otp}'")

        # Click the "Proceed" button and wait for the "custumSearchId" element
        safe_click(driver, (By.XPATH, "//input[@value='Proceed']"), SHORT_WAIT_SECONDS)

        try:
            safe_click(driver, (By.ID, "custumSearchId"), DEFAULT_WAIT_SECONDS)
            return driver  # Return driver if login succeeds
        except Exception as e:
            print(f"Attempt {attempt + 1} failed - Exception: you have entered wrong verification code")
            driver.get("https://www.ireps.gov.in/epsn/anonymSearch.do")
            time.sleep(5)

    return driver  # Return driver if all login attempts fail


# def login(driver, mobile_no):

#     retries = 0
#     while retries < 3:
#         try:
#             driver.refresh()
#             # If the refresh succeeds, break out of the loop
#             break
#         except TimeoutException:
#             print("Timeout exception occurred. Retrying...")
#             retries += 1
#             # Add some delay before retrying to avoid overwhelming the server
#             time.sleep(2)
#     else:
#         # If all retries fail, raise the TimeoutException
#         raise TimeoutException("Exceeded maximum retries. Unable to refresh.")

#     time.sleep(3)

#     # Retry login process up to 3 times
#     for attempt in range(3):
#         try:

#             try:
#                 alert = Alert(driver)
#                 alert.accept()
#             except:
#                 pass

#             driver, ver_code = get_verification(driver)

#             if ver_code is None:
#                 return None

#             otp = load_otp()
#             # Fill in login details and proceed
#             driver.execute_script("document.getElementById('mobileNo').value='" + mobile_no + "'")
#             time.sleep(1)
#             driver.execute_script("document.getElementById('verification').value='" + ver_code + "'")
#             time.sleep(1)
#             driver.execute_script("document.getElementById('otp').value='" + otp + "'")
#             time.sleep(2)
#             driver.find_element("xpath", "//input[@value='Proceed']").click()

#             # WebDriverWait block to wait for the presence of the element with ID "customSearchId"
#             driver.find_element(By.ID, "custumSearchId").click()
#             # If everything is successful, break out of the loop
#             return driver

#         except Exception as e:
#             # Handle other exceptions while clicking 'Custom Search' button
#             print(f"Attempt {attempt + 1} login or Custom Search button  -  Exception") # {e}")
#             driver.get("https://www.ireps.gov.in/epsn/anonymSearch.do")
#             time.sleep(5)

#     return driver



def is_no_result_found_present_in_page(driver):
    """
    Check if the page contains the "No Results Found" message.

    Args:
        driver (WebDriver): The Selenium WebDriver instance for interacting with the page.

    Returns:
        tuple: (bool, driver) where the boolean is True if "No Results Found" is present, False otherwise.
    """
    try:
        # Get the page source using Selenium
        page_source = driver.page_source

        # Parse the HTML content with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the element containing the specific class and style attributes
        result_element = soup.find('td', {'class': 'formLabel', 'style': 'color: #C00000'})

        # Check if the element contains the text "No Results Found"
        if result_element and "No Results Found" in result_element.get_text(strip=True):
            print("No Results Found in the page.\n")
            return True, driver

        return False, driver

    except NoSuchElementException as e:
        print(f"An error occurred while checking the page: {e}")
        return False, driver








def download_pdf(url, pdf_file_path, retries=PDF_DOWNLOAD_RETRIES):
    os.makedirs(os.path.dirname(pdf_file_path), exist_ok=True)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    }

    for i in range(retries):
        tmp_file_path = f"{pdf_file_path}.download"
        downloaded_bytes = 0
        try:
            with requests.get(url, headers=headers, timeout=PDF_DOWNLOAD_TIMEOUT, stream=True) as response:
                response.raise_for_status()

                with open(tmp_file_path, 'wb') as output_file:
                    for chunk in response.iter_content(chunk_size=PDF_DOWNLOAD_CHUNK_SIZE):
                        if not chunk:
                            continue
                        output_file.write(chunk)
                        downloaded_bytes += len(chunk)

            if downloaded_bytes == 0:
                raise ValueError("Downloaded PDF is empty")

            os.replace(tmp_file_path, pdf_file_path)
            print("Download successful")
            return True
        except Exception as e:
            if os.path.exists(tmp_file_path):
                os.remove(tmp_file_path)
            print(f"Attempt ({i + 1}/{retries}) failed. Retrying download: {e}")
            time.sleep(min(2 ** i, 10))

    print("All download attempts failed.")
    return False



def getpdfdata(pdf_file_path):
    default_values = ("", "", "", "", "", "", "", "", "", "", "", "")

    try:
        with pdfplumber.open(pdf_file_path) as pdf:
            if not pdf.pages:
                print("PDF has no pages.")
                return default_values

            first_page = pdf.pages[0]
            pdf_heading_text = first_page.extract_text() or ""
            dept_rly = pdf_heading_text.split('\n')[0].strip() if pdf_heading_text else ""

            tables = first_page.extract_tables() or []
            table = tables[0] if tables else []
            table = [[value for value in sublist if value is not None] for sublist in table]
            new_table = [row[i * 2:i * 2 + 2] for row in table for i in range(len(row) // 2)] + [row for row in table if len(row) <= 2]
            new_table = [[cell.replace('\n', ' ') for cell in row] for row in new_table]

            extracted_info = {}
            keys_to_extract = [
                'Name of Work', 'Bidding type', 'Tender Type', 'Bidding System',
                'Tender Closing Date Time', 'Date Time Of Uploading Tender',
                'Pre-Bid Conference Date Time', 'Advertised Value',
                'Earnest Money (Rs.)', 'Contract Type'
            ]

            for item in new_table:
                if len(item) > 1 and item[0] in keys_to_extract:
                    extracted_info[item[0]] = item[1]

            page_text = pdf_heading_text
            tender_no = ""
            tender_no_index = page_text.find("Tender No:")
            if tender_no_index != -1:
                tender_no_text = page_text[tender_no_index + len("Tender No:"):].strip()
                tender_no = tender_no_text.split('\n')[0].strip() if '\n' in tender_no_text else tender_no_text
                if len(tender_no) > 36:
                    tender_no = tender_no[:-36]

            name_of_work = extracted_info.get('Name of Work', '')
            bidding_type = extracted_info.get('Bidding type', '')
            tender_type = extracted_info.get('Tender Type', '')
            bidding_system = extracted_info.get('Bidding System', '')
            tender_closing_date_time = extracted_info.get('Tender Closing Date Time', '')
            date_time_of_uploading_tender = extracted_info.get('Date Time Of Uploading Tender', '')
            pre_bid_conference_date_time = extracted_info.get('Pre-Bid Conference Date Time', '')
            advertised_value = extracted_info.get('Advertised Value', '')
            earnest_money = extracted_info.get('Earnest Money (Rs.)', '')
            contract_type = extracted_info.get('Contract Type', '')

            # print("Department:", dept_rly)
            # print("Tender No:", tender_no)
            # print("Name of Work:", name_of_work)
            # print("Bidding Type:", bidding_type)
            # print("Tender Type:", tender_type)
            # print("Bidding System:", bidding_system)
            # print("Tender Closing Date and Time:", tender_closing_date_time)
            # print("Date and Time of Uploading Tender:", date_time_of_uploading_tender)
            # print("Pre-Bid Conference Date and Time:", pre_bid_conference_date_time)
            # print("Advertised Value:", advertised_value)
            # print("Earnest Money:", earnest_money)
            # print("Contract Type:", contract_type)

            return dept_rly, tender_no, name_of_work, bidding_type, tender_type, bidding_system, tender_closing_date_time, date_time_of_uploading_tender, pre_bid_conference_date_time, advertised_value, earnest_money, contract_type
    except Exception as exc:
        print(f"Unable to read tender PDF data: {exc}")
        return default_values



def select_organization(driver, org_number):
    """Select the requested organization on the IREPS search page."""
    for _ in range(3):  # Try the action three times
        try:
            dropdown_element = WebDriverWait(driver, DEFAULT_WAIT_SECONDS).until(
                EC.presence_of_element_located((By.ID, "organization"))
            )
            select = Select(dropdown_element)
            select.select_by_value(org_number)
            time.sleep(1)

            if dropdown_element:
                driver.execute_script("document.getElementById('organization').value='" + org_number + "'")
            else:
                print("Element with ID 'organization' not found")
            return True

        except Exception:
            print("organization dropdown  -  Exception")
            driver.refresh()
            time.sleep(2)

    return False



def get_zone_options(driver):
    """Return all configured zone dropdown options for the selected organization."""
    time.sleep(1)
    railway_zone_element = WebDriverWait(driver, DEFAULT_WAIT_SECONDS).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='railwayZone']"))
    )
    railway_zone_dropdown = Select(railway_zone_element)

    options_dict = {}
    for option in railway_zone_dropdown.options:
        value = option.get_attribute("value")
        text = option.get_attribute("innerText")
        options_dict[value] = text

    zone_options = [
        (zone_number, zone)
        for zone_number, zone in options_dict.items()
        if zone not in skip_zones
    ]

    print("\n--------- ZONE LIST ---------")
    for _, zone in zone_options:
        print(zone)

    return zone_options



def scrape_zone(driver, org_number, org_name, program_file_dir, zone_number, zone):
    """Scrape one organization zone with the supplied Selenium driver."""
    last_tender = False
    print(f"\nScraping ZONE -> {zone}")
    print("----------------")


    for _ in range(3):  # Try the action three times
        try:
            # filling search criteria
            time.sleep(3)
            driver.execute_script("document.getElementById('organization').value='"+ org_number +"'")
            time.sleep(3)
            driver.execute_script("document.getElementById('workArea').value='WT'") # works
            time.sleep(3)
            driver.execute_script("document.getElementById('railwayZone').value='"+ zone_number +"'")
            time.sleep(3)
            driver.execute_script("document.getElementById('tenderType').value=2") # open
            time.sleep(3)
            driver.execute_script("document.getElementById('tenderStage').value=1") # published
            time.sleep(3)
            driver.execute_script("document.getElementsByName('selectDate')[0].value = 'TENDER_OPENING_DATE'") # Tender Closing Date
            # Get the current date
            current_date = datetime.datetime.now()
            # Add four months to the current date
            four_months_later = current_date + relativedelta(months=4)
            # Format the date as a string (optional)
            formatted_date = four_months_later.strftime("%d/%m/%Y")
            driver.execute_script("document.getElementById('ddmmyyDateformat2').value='" + formatted_date + "'")
            time.sleep(0.5)
            # driver.find_element(By.XPATH, "//input[@value='Show Results']").click()
            xpath = "//input[@value='Show Results']"
            element = driver.find_element(By.XPATH, xpath)
            element.click()

            # If everything is successful, break out of the loop
            break

        except Exception as e:
            # Handle other exceptions while clicking 'Custom Search' button
            print(f"Show Results button  -  Exception")
            driver.refresh()
            time.sleep(2)

    result, driver = is_no_result_found_present_in_page(driver)

    packages = packaging()
    if result:
        return False

    try:
        tender_count = 0
        time.sleep(3.5)

        # Get the page source and parse it with BeautifulSoup
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Find all 'b' tags within 'tr' tags
        b_tags = soup.find_all('b')

        # Iterate over the 'b' tags
        for i in range(len(b_tags)):
            # If the 'b' tag contains "Tender search results"
            if "Tender search results" in b_tags[i].text:
                # If there is a next 'b' tag, print it
                if i + 1 < len(b_tags):
                    tender_count = b_tags[i + 1].text
                    if int(tender_count) < 1 :  # Check if tender_count exists but is None
                        print("Unable to get tender search result, Variable is None and has no value assigned. ", tender_count)
                        continue
                    print("Tender search results ", tender_count)
                break

    except Exception as e:
        print(f"An error occurred: Tender search results", e)
        return False

    # # Get the page source and parse it with BeautifulSoup
    # soup = BeautifulSoup(driver.page_source, 'html.parser')
    # # Find all 'b' tags within 'tr' tags
    # b_tags = soup.find_all('b')
    # # Iterate over the 'b' tags
    # for i in range(len(b_tags)):
    #     # If the 'b' tag contains "Tender search results"
    #     if "Tender search results" in b_tags[i].text:
    #         # If there is a next 'b' tag, print it
    #         if i+1 < len(b_tags):
    #             tender_count = b_tags[i+1].text
    #             print("Tender search results ", tender_count)
    #         break


    # Get the current date and time
    current_date = datetime.datetime.now()
    # Format the date and time string
    fname = current_date.strftime("%d-%m-%Y %H_%M_%S")
    # Create a file name using the value and the formatted date and time
    file_name = f'{zone}_{fname}.xlsx'
    org_folder_path = os.path.join(program_file_dir, org_name)
    os.makedirs(org_folder_path, exist_ok=True)
    # Create a file path for the new Excel workbook
    file_path = os.path.join(org_folder_path, file_name)
    # Create a new Excel workbook and a worksheet within it
    workbook = xlsxwriter.Workbook(file_path, {'constant_memory': True})
    worksheet1 = workbook.add_worksheet("ListOfTenders")
    # Write the column headers in the worksheet
    headers = ["Zone", "Dept.", "Tender No.", "Tender Title", "Type", "Due Date/Time", "Due Days", "Advertised Value", "Doc Link", "Bidding type", "Bidding System", "Date Time Of Uploading Tender", "Pre-Bid Conference Date Time", "Earnest Money (Rs.)", "Contract Type"]
    for index, header in enumerate(headers):
        worksheet1.write(0, index, header)


    # Calculate the number of pages based on the tender count
    tender_count = int(tender_count)
    decimal_number = tender_count / 25
    # Round up to the nearest integer to get the page count
    page_count = math.ceil(decimal_number)
    # Print the number of pages
    print("Pages = ", page_count)


    cnt = 1
    k = 0
    # Loop through all the pages
    while cnt <= page_count:
        time.sleep(3)
        for _ in range(3):  # Try the action three times
            try:
                # If the element is found, you can perform further actions here
                a_tags = driver.find_elements(By.CSS_SELECTOR, "a[onclick]")
                # If everything is successful, break out of the loop
                break
            except Exception as e:
                # Handle other exceptions while clicking 'Custom Search' button
                print(f"a_tags tender link - Exception")
                time.sleep(2)
                continue

        if not a_tags:
            break

        filtered_a_tags = [tag for tag in a_tags if 'postRequestNewWindow(\'/epsn/nitViewAnonyms/rfq/nitPublish.do?' in tag.get_attribute('onclick')]

        # Get the initial window handle
        initial_handle = driver.current_window_handle
        for a_tag in filtered_a_tags:
            k += 1
            print('\r' + "Tender  : " + str(k) + " ", end='')

            try:
                existing_handles = driver.window_handles
                a_tag.click()
                tender_window = wait_for_new_window(driver, existing_handles)
                driver.switch_to.window(tender_window)
            except Exception as e:
                print(f"Result page tender link click - Exception: {e}")
                close_extra_windows(driver, initial_handle)
                time.sleep(2)
                continue


            # # checkpoint
            # temp_url = driver.current_url
            # print(" >> ", temp_url)


            try:
                xpath = "//a[contains(text(), 'Download Tender Doc. (Pdf)')]"
                existing_handles = driver.window_handles
                safe_click(driver, (By.XPATH, xpath), DEFAULT_WAIT_SECONDS)
                pdf_window = wait_for_new_window(driver, existing_handles)
                driver.switch_to.window(pdf_window)
            except Exception as e:
                # Handle other exceptions while clicking 'Custom Search' button
                print(f"Download Tender Doc. (Pdf) button - Exception: {e}")
                close_extra_windows(driver, initial_handle)
                time.sleep(2)
                continue

            handles = driver.window_handles


            # # Define a function to wait for the page to fully load
            # def page_fully_loaded(driver):
            #     return driver.execute_script("return document.readyState") == "complete"

            # # Wait for the page to fully load
            # WebDriverWait(driver, 10).until(page_fully_loaded)

            pdf_url = driver.current_url
            print(" ", pdf_url)

            # # Execute JavaScript to get the current window's URL
            # window_url = driver.execute_script("return window.location.href;")
            # url_pattern = re.compile(r'^https:\/\/www\.ireps\.gov\.in\/ireps\/works\/pdfdocs\/.*\.pdf$')

            # while True:
            #     if url_pattern.match(pdf_url):
            #         print("URL is valid. ", pdf_url)
            #         break
            #     else:
            #         print("URL is not valid.")
            #         time.sleep(0.25)
            #         pdf_url = driver.current_url
            #         continue

            if pdf_url.lower().endswith(".pdf"):
                pdf_file_path = os.path.join(temp_dir_path, f"tender_{threading.get_ident()}_{uuid.uuid4().hex}.pdf")
                if not download_pdf(pdf_url, pdf_file_path):
                    close_extra_windows(driver, initial_handle)
                    continue
            else:
                close_extra_windows(driver, initial_handle)
                continue

            dept_rly, tender_no, name_of_work, bidding_type, tender_type, bidding_system, tender_closing_date_time, date_time_of_uploading_tender, pre_bid_conference_date_time, advertised_value, earnest_money, contract_type = getpdfdata(pdf_file_path)
            if os.path.exists(pdf_file_path):
                os.remove(pdf_file_path)

            try:
                # Calculate due_days
                closing_datetime = datetime.datetime.strptime(tender_closing_date_time, '%d/%m/%Y %H:%M')
                uploading_datetime = datetime.datetime.strptime(date_time_of_uploading_tender, '%d/%m/%Y %H:%M')
                due_days = (closing_datetime - uploading_datetime).days
            except ValueError as e:
                due_days = " "
                pass

            worksheet1.write(k, 0, zone)
            worksheet1.write(k, 1, dept_rly)
            worksheet1.write(k, 2, tender_no)
            worksheet1.write(k, 3, name_of_work)
            worksheet1.write(k, 4, tender_type)
            worksheet1.write(k, 5, tender_closing_date_time)
            worksheet1.write(k, 6, due_days)
            worksheet1.write(k, 7, advertised_value)
            worksheet1.write(k, 8, pdf_url)
            worksheet1.write(k, 9, bidding_type)
            worksheet1.write(k, 10, bidding_system)
            worksheet1.write(k, 11, date_time_of_uploading_tender)
            worksheet1.write(k, 12, pre_bid_conference_date_time)
            worksheet1.write(k, 13, earnest_money)
            worksheet1.write(k, 14, contract_type)

            # Switch back to the initial window
            close_extra_windows(driver, initial_handle)
            gc.collect()

            if k == tender_count:
                last_tender = True
            else:
                last_tender = False

        print("\n")

        if last_tender == True:
            break
        else:

            try:
                xpath = f"//a[text()='{cnt + 1}']"
                for attempt in range(3):
                    try:
                        # Wait up to 5 seconds for the element to be clickable
                        element = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, xpath))
                        )
                        element.click()
                        print(f"Successfully clicked page {cnt + 1} on attempt {attempt + 1}")
                        time.sleep(5)
                        break  # Success!
                    except Exception as e:
                        print(f"Attempt {attempt + 1} failed to click page {cnt + 1}: {e}")
                        time.sleep(2)
                else:
                    print(f"Failed to click page {cnt + 1} after 3 attempts.")

            except Exception as e:
                print(f"Exception while trying to click page {cnt + 1}: {e}")

        if cnt % 10 == 0:
            print("\n")
            try:
                # driver.find_element(By.XPATH, f"//a[font[text()='next']]").click()
                xpath = "//a[font[text()='next']]"
                element = driver.find_element(By.XPATH, xpath)
                element.click()

            except Exception as e:
                print(f"Element with text 'next' not found")
                break

        gc.collect()
        cnt += 1

    # Close the workbook and print a message
    workbook.close()
    print("Zone data Saved.")



def tenders(driver, org_number, org_name, program_file_dir):
    if not select_organization(driver, org_number):
        return False

    zone_options = get_zone_options(driver)
    if org_number == INDIAN_RAILWAY_ORG_NUMBER:
        return scrape_indian_railway_zones_with_workers(org_number, org_name, zone_options)

    mail_trigger = False
    for zone_number, zone in zone_options:
        mail_trigger = scrape_zone(driver, org_number, org_name, program_file_dir, zone_number, zone) or mail_trigger

    return mail_trigger



# extract message from adb device connected
def get_sms_message():
    message = ""
    device_serial = get_current_device_serial()
    command = f'adb -s {device_serial} shell content query --uri content://sms/inbox'
    try:
        output = subprocess.check_output(command, shell=True, encoding='utf-8').strip()
        lines = output.split('\n')
        # print(lines)
        for line in lines:
            # Extract the date from the message
            match_date = re.search(r'date=(\d+)', line)
            if match_date:
                timestamp = int(match_date.group(1)) / 1000  # Convert from milliseconds to seconds
                message_date = datetime.datetime.fromtimestamp(timestamp).date()
                # Compare the message date with today's date
                if message_date == datetime.datetime.today().date():  # Use datetime.today() instead of datetime.date.today()
                    # print(line)
                    if 'IREPS' in line:
                        match_otp = re.search(r'body=(\d{6})', line)
                        if match_otp:
                            six_digits = match_otp.group(1)
                            print(f"OTP: {six_digits}")

                            # print(type(message_date))
                            # countdown_timer(10)

                            # Read the JSON file
                            with open(config_file_path, 'r') as file:
                                config_data = json.load(file)

                            # Update the OTP date and OTP
                            config_data['otp_date'] = message_date.strftime("%Y-%m-%d")
                            config_data['otp'] = six_digits

                            # Write the updated data back to the file
                            with open(config_file_path, 'w') as file:
                                json.dump(config_data, file, indent=4)

                            # with open(otp_full_path, "w") as file:
                            #     file.write(f"Date: {message_date}\n")
                            #     file.write(f"OTP: {six_digits}\n")
                            return False
                        else:
                            print("Pattern not found.")
                    else:
                        message = "Messages not yet received; retring to access the OTP"
        print(message)
        return True
    except subprocess.CalledProcessError as e:
        print(f"Error executing command: {e}")
        return True




# generate OTP
def generate_otp(driver, mobile_no):
    driver.refresh()
    time.sleep(3)

    # print("Current Mobile No. :", mobile_no)
    driver, Verification_code = get_verification(driver)
    # mobile_no = input("Enter 10 digit Mobile No: ")
    driver.execute_script("document.getElementById('mobileNo').value='" + mobile_no + "'")

    driver.execute_script("document.getElementById('verification').value='" + Verification_code + "'")


    driver.find_element("xpath", "//input[@value='Get OTP']").click()
    time.sleep(3)

    try:
        # Check if an alert is present
        alert = driver.switch_to.alert
        print("Alert Text:", alert.text)
        if alert.text == "you have entered wrong verification code.":
            alert.accept()
            generate_otp(driver, mobile_no)

        alert.accept()  # Close the alert (Accept/Dismiss)
    except NoAlertPresentException:
        print("No alert present after clicking 'Get OTP'")
    return driver




def create_logged_in_driver(mobile_no):
    """Create a Chrome driver, open IREPS, refresh OTP when needed, and log in."""
    with CHROMEDRIVER_INSTALL_LOCK:
        chromedriver_autoinstaller.install()  # Check if the current version of chromedriver exists
                                            # and if it doesn't exist, download it automatically,
                                            # then add chromedriver to path

    options = Options()
    options.add_argument("--disable-application-cache")  # Disable application cache
    options.add_argument('--ignore-certificate-errors')
    if browser == "0":
        options.add_argument("--headless")

    # Keep Chrome lightweight during long scraping runs.
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-gpu")
    options.add_argument("--disk-cache-size=1")
    options.add_argument("--media-cache-size=1")
    options.add_argument("--log-level=3")

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    driver.implicitly_wait(10) # Wait for a few seconds to see the results

    url = "https://www.ireps.gov.in/epsn/anonymSearch.do"
    while True:
        try:
            driver.get(url)
            break # break the loop if no exception occurs
        except Exception:
            print(f"{url} - url exception")
            print("Retrying...")
            time.sleep(2)

    driver = is_under_maintenance(driver, url)
    print("Current Mobile NO. ", mobile_no)

    if not is_otp_valid():
        with OTP_REFRESH_LOCK:
            if not is_otp_valid():
                driver = generate_otp(driver, mobile_no)
                countdown_timer("generate_otp", 60)

                while get_sms_message():
                    countdown_timer("get_sms_message", 20)
                    get_sms_message()

    return login(driver, mobile_no)



def tenders_main(org_number, org_name, mobile_no, program_files_dir):
    mail_triger = False
    driver = None
    try:
        driver = create_logged_in_driver(mobile_no)
        mail_triger = tenders(driver, org_number, org_name, program_files_dir)
    finally:
        print(f"\n\nExeting.... From {org_name}.\n\n")
        if driver is not None:
            try:
                driver.quit()
            except Exception as exc:
                print(f"Unable to quit Chrome cleanly: {exc}")
        gc.collect()

    return mail_triger


def get_configured_zone_worker_count(total_zones):
    """Return the safe Indian Railway zone worker count from Configration.json."""
    if total_zones <= 0:
        return 0

    try:
        configured_workers = int(max_zone_workers)
    except (TypeError, ValueError):
        configured_workers = 1

    configured_workers = max(1, configured_workers)

    if str(captcha_manual_input) == "1" and configured_workers > 1:
        print("Manual CAPTCHA input is enabled; forcing Indian Railway zone workers to 1.")
        configured_workers = 1

    return min(configured_workers, total_zones)



def scrape_indian_railway_zone_worker(org_number, org_name, zone_option):
    """Open a separate logged-in browser worker for one Indian Railway zone."""
    zone_number, zone = zone_option
    driver = None
    try:
        print(f"Starting Indian Railway zone worker for {zone}.")
        driver = create_logged_in_driver(mobile_no)
        if not select_organization(driver, org_number):
            return False
        return bool(scrape_zone(driver, org_number, org_name, program_files_dir, zone_number, zone))
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception as exc:
                print(f"Unable to quit Indian Railway zone worker Chrome cleanly: {exc}")
        print(f"Finished Indian Railway zone worker for {zone}.")
        gc.collect()



def scrape_indian_railway_zones_with_workers(org_number, org_name, zone_options):
    """Run separate multi-threaded workers for zones of organization 01: Indian Railway."""
    if not zone_options:
        print("No Indian Railway zones configured for scraping.")
        return False

    worker_count = get_configured_zone_worker_count(len(zone_options))
    print(
        f"Starting {worker_count} Indian Railway zone worker(s) "
        f"for {len(zone_options)} zone(s)."
    )

    mail_trigger_main = False
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        future_to_zone = {
            executor.submit(scrape_indian_railway_zone_worker, org_number, org_name, zone_option): zone_option
            for zone_option in zone_options
        }

        for future in as_completed(future_to_zone):
            zone_number, zone = future_to_zone[future]
            try:
                mail_trigger_main = bool(future.result()) or mail_trigger_main
            except Exception as exc:
                print(f"Indian Railway zone worker failed for {zone_number}: {zone}. Error: {exc}")

    print("All Indian Railway zone workers finished.")
    return mail_trigger_main


def get_configured_org_worker_count(total_organizations):
    """Return the safe organization worker count from Configration.json."""
    if total_organizations <= 0:
        return 0

    try:
        configured_workers = int(max_org_workers)
    except (TypeError, ValueError):
        configured_workers = 1

    configured_workers = max(1, configured_workers)

    if str(captcha_manual_input) == "1" and configured_workers > 1:
        print("Manual CAPTCHA input is enabled; forcing organization workers to 1.")
        configured_workers = 1

    return min(configured_workers, total_organizations)



def scrape_organization(orginazation):
    """Scrape one organization and return True when it produced mail-worthy data."""
    org_number, org_name = orginazation
    print(f"\nRunning Orginazation. {org_number}: {org_name}")
    mail_trigger = tenders_main(org_number, org_name, mobile_no, program_files_dir)
    print(f"Finished Orginazation. {org_number}: {org_name}\n")
    return bool(mail_trigger)



def scrape_organizations_with_workers(orginazations):
    """Run organization scraping workers and wait for all of them before returning."""
    if not orginazations:
        print("No organizations configured for scraping.")
        return False

    worker_count = get_configured_org_worker_count(len(orginazations))
    print(f"Starting {worker_count} organization worker(s) for {len(orginazations)} organization(s).")

    mail_trigger_main = False
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        future_to_organization = {
            executor.submit(scrape_organization, orginazation): orginazation
            for orginazation in orginazations
        }

        for future in as_completed(future_to_organization):
            org_number, org_name = future_to_organization[future]
            try:
                mail_trigger_main = bool(future.result()) or mail_trigger_main
            except Exception as exc:
                print(f"Organization worker failed for {org_number}: {org_name}. Error: {exc}")

    print("All organization workers finished. Starting xlsx merge.")
    return mail_trigger_main


def merge_xlsx_files_in_folders(folders, output_directory, program_file_dir):
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
        print(f"Folder '{output_directory}' created successfully.")

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    output_filename = f"merged_IREPS_{timestamp}.xlsx"
    output_file_path = os.path.join(output_directory, output_filename)
    get_date = datetime.datetime.now().strftime("%d/%m/%Y")

    output_workbook = Workbook(write_only=True)
    output_sheet = output_workbook.create_sheet("ListOfTenders")
    header_written = False
    source_files = 0
    total_rows = 0

    try:
        for folder_name in folders:
            folder_path = os.path.join(program_file_dir, folder_name[1])  # Replace 'path_to_root_folder' with the actual root folder path

            if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
                continue

            for file in os.listdir(folder_path):
                if not file.endswith('.xlsx'):
                    continue

                file_path = os.path.join(folder_path, file)
                source_workbook = None
                try:
                    source_workbook = load_workbook(file_path, read_only=True, data_only=True)
                    source_sheet = source_workbook.active
                    rows = source_sheet.iter_rows(values_only=True)
                    source_header = next(rows, None)

                    if source_header is None:
                        continue

                    source_files += 1
                    if not header_written:
                        output_sheet.append(list(source_header) + ['Get Date'])
                        header_written = True

                    for row in rows:
                        if row is None or all(cell is None for cell in row):
                            continue
                        output_sheet.append(list(row) + [get_date])
                        total_rows += 1
                finally:
                    if source_workbook is not None:
                        source_workbook.close()
                    gc.collect()

        if not header_written:
            output_sheet.append(['Get Date'])

        output_workbook.save(output_file_path)
        print(f"Merged {total_rows} rows from {source_files} xlsx file(s).")
    finally:
        output_workbook.close()
        gc.collect()

    return output_file_path

# # This function is use to merge xlsx files by first argument is (list of tuples)
# def merge_xlsx_files_in_folders(folders, output_directory, program_file_dir):
#     if not os.path.exists(output_directory):
#         os.makedirs(output_directory)
#         print(f"Folder '{output_directory}' created successfully.")

#     b = datetime.datetime.now()
#     merged_data = pd.DataFrame()

#     for folder_name in folders:
#         folder_path = os.path.join(program_files_dir, folder_name[1])  # Replace 'path_to_root_folder' with the actual root folder path

#         # Check if the folder exists before attempting to process it
#         if os.path.exists(folder_path) and os.path.isdir(folder_path):
#             files = os.listdir(folder_path)

#             for file in files:
#                 if file.endswith('.xlsx'):
#                     file_path = os.path.join(folder_path, file)
#                     data = pd.read_excel(file_path)
#                     b = datetime.datetime.now()
#                     data['Get Date'] = b.strftime("%d/%m/%Y")
#                     merged_data = pd.concat([merged_data, data], ignore_index=True)

#         # else:
#         #     # Folder not found, so skip this iteration
#         #     print(f"Folder not found: {folder_path}. Skipping.")

#     # Create the timestamp for the filename
#     timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

#     # Output file path and name
#     output_filename = f"merged_IREPS_{timestamp}.xlsx"
#     output_file_path = os.path.join(output_directory, output_filename)

#     # Save the merged data to the output file
#     merged_data.to_excel(output_file_path, index=False)

#     return output_file_path









# # Load dump location from JSON file
# def load_email_flag():
#     try:
#         with open(config_file_path, "r") as file:
#             data = json.load(file)
#             return data.get("email_flag")
#     except FileNotFoundError:
#         return None







# smtp send status
def send_mail(program_file_dir, all_email_ids, master_xlsx_path=None):
    # Check if read_email_ids list has at least one item
    if not all_email_ids:
        print("Email skipped: receiver_emailids is empty.")
        return False

    if not sender_email_id or not sender_email_password:
        print("Email skipped: sender_email_id or sender_email_password is missing in Configration.json.")
        return False

    sender_email = sender_email_id
    receiver_emails = all_email_ids # ["am7059141480@gmail.com", "vmaskara@royalconstruct.com"]
    subject = "IREPS Tender Scraping Report"

    notification_text = "This is an automated notification to inform you that the IREPS tender scraping process has been completed. The master xlsx file is attached. Please verify the results. \n\n"

    # Gather system information
    system_info = (
        f"System: {platform.system()}\n"
        f"Hostname: {socket.gethostname()}\n"
        f"IP Address: {socket.gethostbyname(socket.gethostname())}\n"
        f"Working Directory: {os.getcwd()}\n"
        f"Windows Version: {platform.win32_ver()[1]}\n"
    )
    # Convert dictionary to JSON format
    # json_data = json.dumps(exception_data_Dictionary, indent=4)

    # Concatenate the notification text, system information, and json data
    # email_body = f"{notification_text}\n{system_info}\n\nLog Data:\n{json_data}"
    email_body = f"{notification_text}\n{system_info}"

    # # Remove brackets from JSON data
    # json_object = json.loads(json_data)
    # if isinstance(json_object, list):
    #     json_data = json_object[0]

    # Create a MIMEText object
    msg = MIMEMultipart()
    msg.attach(MIMEText(email_body))

    # Set the sender and recipients
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = subject

    # # Attach the log file if requested
    # if attach_log:
    #     if os.path.exists(log_file_path) and os.path.getsize(log_file_path) > 0:
    #         with open(log_file_path, "rb") as attachment:
    #             part = MIMEBase("application", "octet-stream")
    #             part.set_payload((attachment).read())
    #             encoders.encode_base64(part)
    #             part.add_header("Content-Disposition", "attachment; filename= %s" % log_file_path)
    #             msg.attach(part)
    #     else:
    #         print(f"The file {log_file_path} is empty or does not exist. Not attaching to the email.")
    #         return

    # Attach the newly created master .xlsx file. If an explicit path was not
    # supplied, fall back to the newest .xlsx in the program folder so the
    # email never attaches an older workbook by accident.
    xlsx_file_path = master_xlsx_path
    if xlsx_file_path is None:
        xlsx_folder = program_file_dir
        xlsx_candidates = [
            os.path.join(xlsx_folder, filename)
            for filename in os.listdir(xlsx_folder)
            if filename.endswith(".xlsx")
        ]
        if xlsx_candidates:
            xlsx_file_path = max(xlsx_candidates, key=os.path.getmtime)

    attachment_found = False
    if xlsx_file_path and os.path.exists(xlsx_file_path) and xlsx_file_path.endswith(".xlsx"):
        filename = os.path.basename(xlsx_file_path)
        with open(xlsx_file_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={filename}")
            msg.attach(part)
        attachment_found = True
        print(f"Attached master xlsx to email: {xlsx_file_path}")

    if not attachment_found:
        print(f"Email warning: master .xlsx attachment not found: {xlsx_file_path}")

    # Establish a connection to the SMTP server
    smtp_server = "smtp.office365.com"
    port = 587

    try:
        with smtplib.SMTP(smtp_server, port, timeout=30) as server:
            server.starttls()
            server.login(sender_email_id, sender_email_password)
            server.sendmail(sender_email, receiver_emails, msg.as_string())
    except smtplib.SMTPAuthenticationError as e:
        print(
            "Email could not be sent: SMTP authentication failed. "
            "Verify sender_email_id/sender_email_password in Configration.json, "
            "use the mailbox app password if MFA is enabled, and confirm SMTP AUTH is enabled for the mailbox. "
            f"Details: {e}"
        )
        return False
    except (smtplib.SMTPException, OSError) as e:
        print(f"Email could not be sent, but scraping output was saved successfully. Error: {e}")
        return False

    print("\nEmail sent successfully!")
    return True









class Tee:
    def __init__(self, *files):
        self.files = files
        self._lock = threading.Lock()

    def write(self, text):
        if not text:
            return

        with self._lock:
            for file in self.files:
                file.write(text)
            if "\n" in text:
                self.flush()

    def flush(self):
        for file in self.files:
            try:
                file.flush()
            except AttributeError:
                pass  # Ignore AttributeError when file is closed



def log_to_file(filename):
    # Create a filename based on current date and time
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_filename = f"{filename}_{timestamp}.txt"
    full_filename = os.path.join(script_dir_path, program_files_dir, "consolelog", new_filename)
    # print(full_filename)
    # time.sleep(1000)
    packge = packaging()


    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(line_buffering=True, write_through=True)

    # Open the log file in line-buffered mode so worker output is written immediately.
    log_file = open(full_filename, 'a', buffering=1, encoding='utf-8')
    original_stdout = sys.stdout
    original_stderr = sys.stderr

    # Redirect stdout/stderr to the log file and console simultaneously.
    realtime_log = Tee(original_stdout, log_file)
    sys.stdout = realtime_log
    sys.stderr = Tee(original_stderr, log_file)

    try:


        # Your main script logic goes here
        print(f"Starting script execution at {timestamp}\n")

        # Record the start time
        start_time = time.time()








        def f2():



            # opening message
            print("\n|-----| Welcome to the IREPS Scraping system! |-----|\n")


            # # Read the JSON file & Update the email_flag
            # with open(config_file_path, 'r') as file:
            #     config_data = json.load(file)
            # config_data['email_flag'] = "False"
            # # config_data['otp_file_location'] = config_file_path

            # # Write the updated data back to the file
            # with open(config_file_path, 'w') as file:
            #     json.dump(config_data, file, indent=4)

            # Deleting xlsx files
            delete_xlsx_files(program_files_dir)


            # Deleting each Folder by the name of Orgination from "Program_Files (ireps_tender_scraping)"
            for orginazation in organizations:
                file_path = os.path.join(script_dir_path, program_files_dir, orginazation[1])
                delete_folder(file_path)



            os.makedirs(temp_dir_path, exist_ok=True)

            if adb_value == "1":
                subprocess.check_output(['adb', 'connect', adb_device_ip]).decode('utf-8').splitlines()
                adb = is_android_device_connected(log_file)
                if adb == False:
                    message = "Please connect an Android device with USB debugging enabled."
                    subject = "IREPS No Android device found."
                    no_adb_mail(subject, message, receiver_emailids)
            else:
                adb = True

            mail_triger_main = False
            if adb:
                # Run Orginazation scraping in parallel and wait for all workers
                # before creating the master xlsx files.
                mail_triger_main = scrape_organizations_with_workers(organizations)


            # master xlsx output directory
            output_directory1 = dump_location  # Replace with the first output directory path
            output_directory2 = program_files_dir   # Replace with the second output directory path
            # creating master xlsx
            output_file_path = merge_xlsx_files_in_folders(organizations, output_directory1, program_files_dir)
            print("Master xlsx saved in ", output_file_path)
            time.sleep(2)
            output_file_path2 = merge_xlsx_files_in_folders(organizations, output_directory2, program_files_dir)
            print("Master xlsx saved in ", output_file_path2)
            time.sleep(2)




            # Final E-Mail with Master xlsx and (log file if exist)
            # mail_triger = load_email_flag()
            # if mail_triger == "True":
            #     if os.path.exists(log_file_path) and os.path.getsize(log_file_path) > 0:
            #         print(program_file_dir, all_email_ids, log_file_path)
            #         send_mail(program_file_dir, all_email_ids, log_file_path, attach_log=True)
            #     else:
            #         print(program_file_dir, all_email_ids, log_file_path)
            #         send_mail(program_file_dir, all_email_ids, log_file_path, attach_log=False)

            if os.path.exists(output_file_path2):
                print(program_files_dir, receiver_emailids, output_file_path2)
                send_mail(program_files_dir, receiver_emailids, output_file_path2)
            else:
                print(f"Email skipped: master xlsx was not saved at {output_file_path2}.")








        def f1():
            for _ in range(3):
                if check_internet_connection():
                    f2()
                    countdown_timer("Exeting in ... ", 10)
                    break
            sys.exit()  # This will exit the Python interpreter

        f1()



        # Record the end time
        end_time = time.time()
        # Calculate the total time taken
        total_time = end_time - start_time
        # Convert seconds to hours, minutes, and seconds
        hours, remainder = divmod(total_time, 3600)
        minutes, seconds = divmod(remainder, 60)
        print(f"Total time taken by the Program: {int(hours)} hours, {int(minutes)} minutes, {int(seconds)} seconds")







        print("Exiting... from tenders download program.")
        # Example: Simulate an exception
        # Uncomment the following line to simulate an exception
        # raise ValueError("This is a simulated exception")

        print("\nScript execution completed.")

    except Exception as e:
        # Print the exception details to both console and error log file
        error_message = f"Exception occurred during script execution: {e}"
        print(error_message, file=sys.stderr)
        log_file.write(error_message + "\n")  # Write the error message to log file

    finally:
        sys.stdout.flush()
        sys.stderr.flush()
        # Restore the original streams
        sys.stdout = original_stdout
        sys.stderr = original_stderr
        log_file.close()              # Close the log file

if __name__ == "__main__":
    create_folder_if_not_exists(os.path.join(script_dir_path, program_files_dir, "consolelog"))
    log_to_file("ireps-tenders-output_log")
    sys.exit()
